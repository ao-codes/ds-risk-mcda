import os
import pathlib
import sys
from abc import abstractmethod
from typing import Any

import numpy as np
import numpy.typing as npt
import openpyxl
import pandas as pd
import yaml
from openpyxl.styles import PatternFill

from ..util import DsRiskMcdaClass


class Assessment(DsRiskMcdaClass):
    @abstractmethod
    def prepare_excel_assessment(self, output_file_path: str) -> None:
        ...

    @classmethod
    @abstractmethod
    def read_excel_assessment(cls, assessment_file_path: str) -> Any:
        ...

    @abstractmethod
    def _check_input(self) -> None:
        ...

    def _finalize_excel_sheet(self, prompt: str, file_path: str) -> None:
        wb = openpyxl.load_workbook(file_path)
        if "Sheet" in wb.sheetnames:
            print(f"{prompt} deleting unnecessary sheet")
            wb.remove(wb["Sheet"])
            wb.save(file_path)
        print(f"{prompt} excel assessment file saved")

    def __remove_old_file_if_existent(self, output_file_path: str) -> None:
        file_path = pathlib.Path(output_file_path)
        if file_path.exists():
            os.remove(file_path)

    def _write_input_data_to_assessment_file(self, list_of_input_dfs: list[pd.DataFrame], output_file_path: str, prefix_sheet_name: str) -> None:
        self.__remove_old_file_if_existent(output_file_path)
        self.__create_file_if_not_existent(output_file_path)
        writer = pd.ExcelWriter(output_file_path, engine="openpyxl", if_sheet_exists="replace", mode="a")
        for i, df in enumerate(list_of_input_dfs):
            sheet_name = prefix_sheet_name + "_" + str(i)
            df.to_excel(writer, sheet_name=sheet_name)
        writer.close()

    @classmethod
    def create_empty_pcm_from_list(cls, input_list: list[str]) -> pd.DataFrame:
        pcm = pd.DataFrame(index=input_list, columns=input_list)
        return pcm

    def __create_file_if_not_existent(self, output_file_path: str) -> None:
        file_exists = os.path.isfile(output_file_path)
        if not file_exists:
            self.__create_new_excel_file(output_file_path)

    def _write_empty_pcm_to_excel_file(self, pcm: pd.DataFrame, output_file_path: str, sheet_name: str, fill_color: str) -> None:
        if len(sheet_name) > 31:
            raise ValueError(f"Please choose a shorter name for '{sheet_name}'. It is recommended to use IDs.")
        writer = pd.ExcelWriter(output_file_path, engine="openpyxl", if_sheet_exists="replace", mode="a")
        pcm.to_excel(writer, sheet_name=sheet_name)
        self.__autofit_columns(writer, sheet_name)
        self.__tint_cells(writer, sheet_name, fill_color)
        writer.close()

    @classmethod
    def complete_pcm(cls, pcm: npt.NDArray[np.float64]) -> npt.NDArray[np.float64]:
        if not isinstance(pcm, np.ndarray):
            raise ValueError("Input is not a Numpy Array!")
        if pd.isnull(pcm).all():
            raise ValueError("Input PCM is empty!")
        for i, row in enumerate(pcm):
            for j, item in enumerate(row):
                if i == j:
                    pcm[i, j] = 1.0
                elif np.isnan(item):
                    pcm[i, j] = 1 / pcm[j, i]
        return pcm

    @classmethod
    def fill_pcm_df(cls, pcm: npt.NDArray[np.float64], df: pd.DataFrame) -> pd.DataFrame:
        for i, row in enumerate(pcm):
            for j, item in enumerate(row):
                df.iloc[i, j] = item
        return df

    @classmethod
    def get_hierarchy_lists(cls, hierarchy_df: pd.DataFrame) -> list[tuple[None, list[str]]] | list[tuple[str, list[str]]]:
        result = []
        first_level = []
        for _, row in hierarchy_df.iterrows():
            if pd.isna(row.iloc[1]):
                first_level.append(row.iloc[0])
        result.append((None, first_level))
        for parent in set(hierarchy_df.iloc[:, 1]):
            if not pd.isna(parent):
                level_elements = []
                for _, row in hierarchy_df.iterrows():
                    if row.iloc[1] == parent:
                        level_elements.append(row.iloc[0])
                result.append((parent, level_elements))
        return result

    def __create_new_excel_file(self, output_file_path: str) -> None:
        wb = openpyxl.Workbook()
        wb.save(output_file_path)

    def __autofit_columns(self, writer: pd.ExcelWriter, sheet_name: str) -> None:
        worksheet = writer.sheets[sheet_name]
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2

    def __tint_cells(self, writer: pd.ExcelWriter, sheet: str, fill_color: str) -> None:
        worksheet = writer.sheets[sheet]
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        for i, row in enumerate(worksheet.iter_rows(min_row=2, min_col=2)):
            for j, cell in enumerate(row):
                cell.number_format = "0.00"
                if i >= j:
                    cell.fill = fill

    @classmethod
    def _check_excel_sheet_names(cls, excel_sheets: dict[str, pd.DataFrame], allowed_sheet_prefixes: list[str]) -> None:
        wrong_sheets = []
        for sheet in excel_sheets:
            is_prefix = False
            for prefix in allowed_sheet_prefixes:
                if sheet.startswith(prefix):
                    is_prefix = True
            if not is_prefix:
                wrong_sheets.append(sheet)
        if len(wrong_sheets) > 0:
            raise ValueError(
                f"The given excel file seems to be no the right one, because there are sheets called {wrong_sheets}. The allowed sheet prefixes are: {allowed_sheet_prefixes}"
            )


class ATAssessment(Assessment):
    """
    A class for creating and reading AHP-TOPSIS assessments.

    Parameters
    ----------
        - criteria_data (dict[str, list[str]] | pd.DataFrame): Data containing information about
          criteria.
        - risks (dict[str, list[str]] | pd.DataFrame): Data containing information about risks.

    Notes
    -----
        criteria_data must be data (in the form of a dict oder a DataFrame) which has 3 columns in
        the following order: column 1: criteria (names or IDs), column 2: parents (parent of the criterion
        if required), column 3: criterion type ("max" or "min").

        risks must be data (in the form of a dict oder a DataFrame) which has 1 column.

    Example
    -------
        criteria_data = {"Criterion": ["S_K", "S", "O", "D", "S_Q"], "CriterionParent": ["S", None,
        None, None, "S"], "CriterionType": ["max", "max", "max", "min", "max"]}

        risks = {"Risks": ["R_01", "R_02", "R_03", "R_04", "R_05", "R_06", "R_07", "R_08", "R_09",
        "R_10"]}

        assessment = ATAssessment(criteria_data, risks)
    """

    def __init__(self, criteria_data: dict[str, list[str]] | pd.DataFrame, risks: dict[str, list[str]] | pd.DataFrame) -> None:
        self.__promt = "AHP-TOPSIS-Assessment:"
        self.__criteria = self._ensure_input_parameter(criteria_data)
        self.__risks = self._ensure_input_parameter(risks)
        self._check_input()
        with open(self._config_file_path, "r", encoding="utf-8") as config_file:
            config = yaml.safe_load(config_file)
            base_level_name = config["ahp_topsis_output_xlsx"]["base_level_name"]
            self.__pcm_sheet_prefix = config["ahp_topsis_output_xlsx"]["pcm_sheet_prefix"]
            self.__sheet_name_level_0 = self.__pcm_sheet_prefix + "_" + base_level_name
            self.__sheet_name_topsis = config["ahp_topsis_output_xlsx"]["sheet_name_topsis"]
            self.__fill_color = config["ahp_topsis_output_xlsx"]["fill_color"]
            self.__input_data_prefix = config["ahp_topsis_output_xlsx"]["input_data_prefix"]

    @property
    def _prompt(self) -> str:
        return self.__promt

    def _check_input(self) -> None:
        if len(self.__criteria.columns) < 3:
            raise ValueError("Incomplete Criteria! Please specifiy a DataFrame (or Dictionary) with 3 columns: criteria, criteria parents and criteria types.")
        risks = self.__risks[self.__risks.columns[0]]
        if not risks.is_unique:
            raise ValueError("Risks have to be unique! Please check input data.")
        criteria = self.__criteria[self.__criteria.columns[0]]
        if not criteria.is_unique:
            raise ValueError("Criteria have to be unique! Please check input data.")
        for element in set(self.__criteria[self.__criteria.columns[1]]):
            elementIsString = element is not None and element is not np.nan
            if elementIsString and element not in self.__criteria[self.__criteria.columns[0]].values:
                raise ValueError(f"You have specified a parent that is no criterion! -> {element}")
        for _, row in self.__criteria.iterrows():
            if row.iloc[0] == row.iloc[1]:
                raise ValueError(f"A parent criterion must not be it's own parent -> '{row.iloc[0]}'")

    def __write_assessment_tables_to_file(self, output_file_path: str) -> None:
        hierarchy_lists = self.get_hierarchy_lists(self.__criteria)
        for hierarchy_level, hierarchy_list in hierarchy_lists:
            pcm = self.create_empty_pcm_from_list(hierarchy_list)
            if pcm.shape == (1, 1):
                raise ValueError("Detected pcm with only 1 element. Please change AHP hierarchy and restart assessment!")
            if hierarchy_level is None:
                sheet_name = self.__sheet_name_level_0
            else:
                sheet_name = self.__pcm_sheet_prefix + "_" + hierarchy_level
            self._write_empty_pcm_to_excel_file(pcm, output_file_path, sheet_name, self.__fill_color)
        topsis_criteria = self.get_topsis_criteria(self.__criteria)
        topsis_df = self.get_topsis_df(topsis_criteria, self.__risks)
        topsis_df.index.rename(None, inplace=True)
        writer = pd.ExcelWriter(output_file_path, engine="openpyxl", if_sheet_exists="replace", mode="a")
        topsis_df.to_excel(writer, sheet_name=self.__sheet_name_topsis)
        writer.close()

    @classmethod
    def get_topsis_criteria(cls, hierarchy_df: pd.DataFrame) -> list[str]:
        result = []
        all_elements = list(hierarchy_df.iloc[:, 0])
        all_parents = set(hierarchy_df.iloc[:, 1])
        for element in all_elements:
            if element not in all_parents:
                result.append(element)
        return result

    @classmethod
    def get_topsis_df(cls, topsis_criteria: list[str], risks: pd.DataFrame) -> pd.DataFrame:
        df = pd.DataFrame(index=risks.iloc[:, 0], columns=topsis_criteria)
        return df

    def prepare_excel_assessment(self, output_file_path: str) -> None:
        """
        Prepares and writes input data to an Excel assessment file.

        Parameters
        ----------
            output_file_path (str): The path to the Excel file to be created or updated.

        Returns
        -------
            None

        Notes
        -----
            This method orchestrates the process of preparing and writing input data to an Excel file
            for assessment purposes. It collects input data from internal criteria and risks DataFrames,
            writes the input data to the specified Excel file, includes assessment tables, and finalizes
            the Excel sheet.

        Example
        -------
            instance.prepare_excel_assessment("output_assessment.xlsx")
        """
        print(f"{self._prompt} writing input data to excel file")
        input_dfs = [self.__criteria, self.__risks]
        self._write_input_data_to_assessment_file(input_dfs, output_file_path, self.__input_data_prefix)
        self.__write_assessment_tables_to_file(output_file_path)
        self._finalize_excel_sheet(self._prompt, output_file_path)

    @classmethod
    def read_excel_assessment(cls, assessment_file_path: str) -> tuple[list[pd.DataFrame], pd.DataFrame, pd.DataFrame]:
        """
        Read an Excel file containing assessment data and extract relevant information.

        Parameters
        ----------
            assessment_file_path (str): The file path of the Excel assessment file.

        Returns
        -------
            tuple: A tuple containing three elements:
                - List of DataFrames: A list of completed Analytic Hierarchy Process (AHP) Pairwise
                  Comparison Matrices (PCMs).
                - DataFrame: A DataFrame representing the TOPSIS decision matrix.
                - DataFrame: A DataFrame representing raw input data.

        Notes
        -----
            The method performs the following tasks: - Reads all sheets from the Excel file. -
            Verifies sheet names based on configured prefixes. - Processes AHP PCMs, TOPSIS data,
            and raw input data from the respective sheets. - Returns a tuple containing lists of AHP
            PCMs, TOPSIS DataFrame, and raw input data DataFrame.

            The completed PCM DataFrames are generated using the AHPAssessment class methods.

        Raises
        ------
            ValueError: If the sheet names in the Excel file do not match the configured prefixes.

        Example
        -------
            ahp_pcms, topsis_df, raw_data_df =
            ATAssessment.read_excel_assessment("/path/to/assessment_file.xlsx")
        """
        excel_sheets = pd.read_excel(assessment_file_path, None, index_col=0)
        ahp_pcms = []
        topsis_df: pd.DataFrame
        with open(ATAssessment._config_file_path, "r", encoding="utf-8") as config_file:
            config = yaml.safe_load(config_file)
            pcm_prefix = config["ahp_topsis_output_xlsx"]["pcm_sheet_prefix"]
            sheet_name_topsis = config["ahp_topsis_output_xlsx"]["sheet_name_topsis"]
            input_data_prefix = config["ahp_topsis_output_xlsx"]["input_data_prefix"]
        allowed_sheet_prefixes = [pcm_prefix, sheet_name_topsis, input_data_prefix]
        cls._check_excel_sheet_names(excel_sheets, allowed_sheet_prefixes)
        for sheet in excel_sheets:
            df = excel_sheets[sheet]
            if str(sheet).startswith(pcm_prefix):
                completed_pcm = cls.complete_pcm(excel_sheets[sheet].to_numpy())
                filled_pcm = cls.fill_pcm_df(completed_pcm, df)
                ahp_pcms.append(filled_pcm)
            elif str(sheet).startswith(sheet_name_topsis):
                topsis_df = excel_sheets[sheet]
            elif str(sheet).startswith(input_data_prefix + "_0"):
                raw_data_df = df.replace(np.nan, None)
        return ahp_pcms, topsis_df, raw_data_df


class AHPAssessment(Assessment):
    """
    A class for creating and reading Analytic Hierarchy Process (AHP) assessments.

    Parameters
    ----------
        - risk_data (dict[str, list[str]] | pd.DataFrame | None): Data containing information about
          risks.
        - additional_level (dict[str, list[str]] | pd.DataFrame | None): Data containing information
          about an additional hierarchy level, if applicable. Default is None.

    Notes
    -----
        risk_data must be data (in the form of a dict oder a DataFrame) which has 2 columns in the
        following order: column 1: risks, column 2: category (e.g. RBS category).

        additional_level must be data (in the form of a dict oder a DataFrame) which has 1 column.

    Example
    -------
        risk_data = {"Risks": ["R-01", "R-02", "R-03", "R-04"], "Categories": ["C-01", "C-01",
        "C-02", "C-02"]}

        additional_level = {"AddLevel": ["P-01", "P-02", "P-03"]}

        assessment = AHPAssessment(risk_data, additional_level)
    """

    def __init__(self, risk_data: dict[str, list[str]] | pd.DataFrame | None, additional_level: dict[str, list[str]] | pd.DataFrame | None = None):
        self.__prompt = "AHP-Assessment:"
        self.__input_df = self._ensure_input_parameter(risk_data)
        self.__additional_level = self._ensure_input_parameter(additional_level)
        self._check_input()
        with open(self._config_file_path, "r", encoding="utf-8") as config_file:
            config = yaml.safe_load(config_file)
            self.__sheet_name_categories = config["ahp_output_xlsx"]["sheet_name_categories"]
            self.__sheet_name_level_1 = config["ahp_output_xlsx"]["sheet_name_level_1"]
            self.__sheet_name_level_2 = config["ahp_output_xlsx"]["sheet_name_level_2"]
            self.__input_data_prefix = config["ahp_output_xlsx"]["input_data_prefix"]
            self.__fill_color = config["ahp_output_xlsx"]["fill_color"]

    @property
    def _prompt(self) -> str:
        return self.__prompt

    def _check_input(self) -> None:
        if self.__input_df is None:
            raise ValueError("Cannot create assessment file when there is no input given in the class constructor.")
        risks = self.__input_df[self.__input_df.columns[0]]
        if not risks.is_unique:
            raise ValueError("Risks have to be unique! Please check input data.")

    def build_risk_and_category_pcms(self, input_df: pd.DataFrame) -> tuple[pd.DataFrame, list[pd.DataFrame]]:
        print(f"{self._prompt} creating empty pcm for AHP base level")
        category_df = self.create_category_pcm(input_df)
        risk_dfs = self.create_risk_dfs(category_df, input_df)
        return category_df, risk_dfs

    def build_additional_level_pcms(self, risks: "pd.Series[str]", additional_level: "pd.Series[str]") -> list[pd.DataFrame]:
        print(f"{self._prompt} creating empty pcm for additional hierarchy elements")
        result_pcms = []
        for risk in risks.to_list():
            print(f"{self._prompt} creating empty pcm for {risk}")
            new_pcm = self.create_empty_pcm_from_list(additional_level.to_list())
            result_pcms.append(new_pcm)
        return result_pcms

    def create_risk_dfs(self, category_df: pd.DataFrame, raw_data_frame: pd.DataFrame) -> list[pd.DataFrame]:
        df_list = []
        risks_column = raw_data_frame.columns[0]
        for category in category_df.index:
            selection = raw_data_frame.iloc[:, 1] == category
            risks_in_category = raw_data_frame[selection][risks_column].to_list()
            if len(risks_in_category) == 1:
                raise ValueError("Detected pcm with only 1 element. Please change AHP hierarchy and restart assessment!")
            else:
                print(f"{self._prompt} creating empty pcm for category '{category}'")
                new_risk_df = self.create_empty_pcm_from_list(risks_in_category)
                df_list.append(new_risk_df)
        return df_list

    def create_category_pcm(self, input_df: pd.DataFrame) -> pd.DataFrame:
        categories_column = input_df.columns[1]
        values = input_df[categories_column].value_counts()
        list_of_risk_categories = values.index.to_list()
        pcm = self.create_empty_pcm_from_list(list_of_risk_categories)
        return pcm

    def write_completed_pcm_to_excel_file(self, assessment_file: str, sheet_name: str) -> None:
        print(f"{self._prompt} Completing pairwise comparison matrix: {sheet_name}")
        pcm_df = pd.read_excel(assessment_file, index_col=0, sheet_name=sheet_name)
        pcm = pcm_df.to_numpy(copy=True)
        try:
            result_pcm = self.complete_pcm(pcm)
        except ValueError:
            print(f"{self._prompt} the pcm in sheet '{sheet_name}' is empty. Please fill out the pcm!")
            return
        filled_df = self.fill_pcm_df(result_pcm, pcm_df)
        writer = pd.ExcelWriter(assessment_file, engine="openpyxl", if_sheet_exists="replace", mode="a")
        filled_df.to_excel(writer, sheet_name=sheet_name)
        self.__autofit_columns(writer, sheet_name)
        self.__tint_cells(writer, sheet_name, self.__fill_color)
        writer.close()

    def complete_excel_pcms(self, assessment_file: str, categories: list[str]) -> None:
        self.write_completed_pcm_to_excel_file(assessment_file, self.__sheet_name_categories)
        for category in categories:
            self.write_completed_pcm_to_excel_file(assessment_file, category)

    def prepare_excel_assessment(self, output_file_path: str) -> None:
        """
        Prepares and writes input data to an Excel assessment file.

        Parameters
        ----------
            output_file_path (str): The path to the Excel file to be created or updated.

        Returns
        -------
            None

        Notes
        -----
            This method orchestrates the process of preparing and writing input data to an Excel file
            for assessment purposes. It collects input data from internal criteria and risks DataFrames,
            writes the input data to the specified Excel file, includes assessment tables, and finalizes
            the Excel sheet.

        Example
        -------
            instance.prepare_excel_assessment("output_assessment.xlsx")
        """
        print(f"{self._prompt} writing input data to excel file")
        input_dfs = [self.__input_df]
        if self.__additional_level is not None:
            input_dfs.append(self.__additional_level)
        self._write_input_data_to_assessment_file(input_dfs, output_file_path, self.__input_data_prefix)
        print(f"{self._prompt} writing pcm for risk categories to excel file")
        category_pcm, risk_pcms = self.build_risk_and_category_pcms(self.__input_df)
        self._write_empty_pcm_to_excel_file(category_pcm, output_file_path, self.__sheet_name_categories, self.__fill_color)
        categories = category_pcm.index
        for i, risk_df in enumerate(risk_pcms):
            if risk_df.shape == (1, 1):
                raise ValueError("Detected pcm with only 1 element. Please change AHP hierarchy and restart assessment!")
            else:
                print(f"{self._prompt} writing pcm for '{categories[i]}' to excel file")
                pcm_sheet_name = self.__sheet_name_level_1 + "_" + categories[i]
                self._write_empty_pcm_to_excel_file(risk_df, output_file_path, pcm_sheet_name, self.__fill_color)
        if self.__additional_level is not None:
            add_level = self.__additional_level.squeeze()
            risks = self.__input_df[self.__input_df.columns[0]]
            add_pcms = self.build_additional_level_pcms(risks, add_level)
            for i, add_pcm in enumerate(add_pcms):
                print(f"{self._prompt} writing additional pcm for '{risks[i]}' to excel file")
                pcm_sheet_name = self.__sheet_name_level_2 + "_" + risks[i]
                self._write_empty_pcm_to_excel_file(add_pcm, output_file_path, pcm_sheet_name, self.__fill_color)
        self._finalize_excel_sheet(self._prompt, output_file_path)

    @classmethod
    def read_excel_assessment(cls, assessment_file_path: str) -> tuple[pd.DataFrame, list[pd.DataFrame], None] | tuple[pd.DataFrame, list[pd.DataFrame], list[pd.DataFrame]]:
        """
        Reads an Excel assessment file, processes the data, and returns the parsed information.

        Parameters
        ----------
            assessment_file_path (str): The file path of the Excel assessment file.

        Returns
        -------
            tuple: A tuple containing processed DataFrames for category, risk, and additional PCMs.
                - pd.DataFrame: The completed PCM DataFrame for categories.
                - list[pd.DataFrame]: List of completed PCM DataFrames for risk.
                - list[pd.DataFrame] or None: List of completed PCM DataFrames for additional criteria, or None if not present.

        Notes
        -----
            - This method assumes a specific structure in the Excel file which was before created by
              the method prepare_excel_assessment.
            - The completed PCM DataFrames are generated using the AHPAssessment class methods.

        Raises
        ------
            ValueError: If the sheet names in the Excel file do not match the configured prefixes.

        Example
        -------
            category_pcm, risk_pcms, add_pcms =
            AHPAssessment.read_excel_assessment("/path/to/assessment_file.xlsx")
        """
        excel_sheets = pd.read_excel(assessment_file_path, None, index_col=0)
        category_pcm: pd.DataFrame
        risk_pcms = []
        add_pcms = []
        with open(AHPAssessment._config_file_path, "r", encoding="utf-8") as config_file:
            config = yaml.safe_load(config_file)
            sheet_name_categories = config["ahp_output_xlsx"]["sheet_name_categories"]
            sheet_name_level_1 = config["ahp_output_xlsx"]["sheet_name_level_1"]
            sheet_name_level_2 = config["ahp_output_xlsx"]["sheet_name_level_2"]
            input_data_prefix = config["ahp_output_xlsx"]["input_data_prefix"]
        allowed_sheet_prefixes = [sheet_name_categories, sheet_name_level_1, sheet_name_level_2, input_data_prefix]
        cls._check_excel_sheet_names(excel_sheets, allowed_sheet_prefixes)
        for sheet in excel_sheets:
            if str(sheet).startswith(input_data_prefix):
                continue
            df = excel_sheets[sheet]
            completed_pcm = cls.complete_pcm(excel_sheets[sheet].to_numpy())
            filled_pcm = cls.fill_pcm_df(completed_pcm, df)
            if str(sheet).startswith(sheet_name_categories):
                category_pcm = filled_pcm
            elif str(sheet).startswith(sheet_name_level_1):
                risk_pcms.append(filled_pcm)
            elif str(sheet).startswith(sheet_name_level_2):
                add_pcms.append(filled_pcm)
        if len(add_pcms) == 0:
            return category_pcm, risk_pcms, None
        return category_pcm, risk_pcms, add_pcms
